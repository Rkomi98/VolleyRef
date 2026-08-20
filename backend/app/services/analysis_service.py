"""Orchestrazione delle Analysis: upload, pipeline, correzioni, export.

Specifica di riferimento: 02_volleyref_backend_prompt.md §4-§20, §44.

La pipeline vera vive in `app.services.extraction_pipeline` (inspect_pdf →
percorso text-layer o raster/OCR → RawObservation → parser → validator →
`Analysis`); questo modulo la orchestra e ne pubblica il risultato.

`build_canned_analysis` (i dati fabbricati del "contratto vivente" di §44)
resta come **fallback esplicito**, per due casi: la pipeline reale disabilitata
via `VOLLEYREF_USE_REAL_PIPELINE=0`, e un'estrazione che solleva un'eccezione su
un PDF non supportato (per esempio un percorso raster senza Tesseract
installato). Il fallback non è mai silenzioso: emette un log di livello ERROR e
inserisce nel risultato il check `pipeline-fallback` con stato `INVALID`, così
che nessuno possa confondere numeri fabbricati con dati letti dal referto.
"""

from __future__ import annotations

import asyncio
import os
import random
import uuid
from pathlib import Path
from typing import Optional

from fastapi import BackgroundTasks

from app.core.logging import get_logger
from app.models.analysis import (
    Analysis,
    AnalysisGlobalStatus,
    AnalysisStatusResponse,
    CreateAnalysisResponse,
    ProcessingStep,
    ProcessingStepId,
    ProcessingStepStatus,
)
from app.models.common import CheckStatus, ErrorCode, ExtractionMethod
from app.models.match import (
    MatchInfo,
    RotationLabel,
    ServiceTurn,
    SetData,
    StartingSix,
    Team,
    ValidationCheck,
    ValidationResult,
)
from app.core.security import resolve_pdf_within_storage
from app.repositories.analysis_repository import AnalysisRecord, AnalysisRepository
from app.services.extraction_pipeline import fallback_check, run_real_pipeline
from app.volleyball.parser import starting_six_field_id

_logger = get_logger(__name__)

# --------------------------------------------------------------------------
# Selettore pipeline reale / fallback canned (backend §44).
# --------------------------------------------------------------------------

REAL_PIPELINE_ENV_VAR = "VOLLEYREF_USE_REAL_PIPELINE"
"""Env var che governa l'uso della pipeline di estrazione reale.

Default: **attiva**. Impostarla a `0`/`false`/`no`/`off` forza il fallback
canned, utile per lavorare sul frontend senza Tesseract installato — ma il
risultato resta marcato `INVALID` con il check `pipeline-fallback`, perché anche
in quel caso i numeri mostrati non vengono dal PDF caricato.
"""

_FALSEY = {"0", "false", "no", "off"}


def real_pipeline_enabled() -> bool:
    raw = os.environ.get(REAL_PIPELINE_ENV_VAR)
    if raw is None:
        return True
    return raw.strip().lower() not in _FALSEY


# --------------------------------------------------------------------------
# Errori di dominio — mappati sulle risposte ErrorEnvelope da main.py.
# --------------------------------------------------------------------------


class AnalysisServiceError(Exception):
    """Base per gli errori di dominio del servizio Analysis.

    `code` è un `ErrorCode` (app/models/common.py); main.py lo usa per
    scegliere lo status HTTP e costruire l'ErrorEnvelope.
    """

    code: ErrorCode = ErrorCode.INTERNAL_ERROR

    def __init__(self, message: str, details: Optional[dict] = None) -> None:
        self.message = message
        self.details = details or {}
        super().__init__(message)


class InvalidFileError(AnalysisServiceError):
    code = ErrorCode.INVALID_FILE


class UnsupportedPdfError(AnalysisServiceError):
    code = ErrorCode.UNSUPPORTED_PDF


class AnalysisNotFoundError(AnalysisServiceError):
    code = ErrorCode.ANALYSIS_NOT_FOUND


class InvalidFieldValueError(AnalysisServiceError):
    code = ErrorCode.INVALID_FIELD_VALUE


class ExportFailedError(AnalysisServiceError):
    code = ErrorCode.EXPORT_FAILED


# --------------------------------------------------------------------------
# Pipeline mock: 5 step, nell'ordine esatto di ProcessingStepId (backend §6).
# --------------------------------------------------------------------------

_STEP_ORDER: list[ProcessingStepId] = [
    ProcessingStepId.READ_DOCUMENT,
    ProcessingStepId.DETECT_SETS,
    ProcessingStepId.EXTRACT_STARTING_SIX,
    ProcessingStepId.EXTRACT_SERVICE_TURNS,
    ProcessingStepId.VALIDATE,
]

_STEP_DELAY_SECONDS = 0.3


def _steps_snapshot(
    completed_up_to: int, processing_index: Optional[int]
) -> list[dict]:
    """Costruisce la lista di step con lo stato corretto per ognuno.

    `completed_up_to` è l'indice (escluso) fino al quale gli step sono
    COMPLETED; `processing_index`, se presente, è l'indice dello step
    attualmente PROCESSING. Gli altri restano PENDING.
    """

    steps: list[dict] = []
    for i, step_id in enumerate(_STEP_ORDER):
        if i < completed_up_to:
            status = ProcessingStepStatus.COMPLETED.value
        elif processing_index is not None and i == processing_index:
            status = ProcessingStepStatus.PROCESSING.value
        else:
            status = ProcessingStepStatus.PENDING.value
        steps.append({"id": step_id.value, "status": status})
    return steps


def _skeleton_analysis(analysis_id: str) -> dict:
    """Analysis minima e sempre valida, usata come placeholder mentre la
    pipeline (mock) non ha ancora prodotto il risultato finale — evita che
    `GET /analyses/{id}` debba gestire uno stato "non ancora esistente" per
    un id che invece esiste (backend §7)."""

    analysis = Analysis(
        id=analysis_id,
        status=AnalysisGlobalStatus.UPLOADED,
        overall_validation=None,
        match=MatchInfo(
            competition=None,
            match_number=None,
            date=None,
            time=None,
            venue=None,
            team_a=Team(id="team-a", name="—"),
            team_b=Team(id="team-b", name="—"),
            final_result=(0, 0),
        ),
        sets=[],
        source_regions=[],
        validation=ValidationResult(status=CheckStatus.VALID, checks=[]),
    )
    return analysis.model_dump(mode="json")


# --------------------------------------------------------------------------
# Costruzione dell'Analysis CANNED (fixture ISUZU CEREA VR vs ROTHOBLAAS
# VOLANO TN — backend §28). Set 1 usa i valori reali documentati; i set 2-4 e
# i turni di servizio sono fabbricati ma internamente coerenti.
# --------------------------------------------------------------------------

_TEAM_A_ID = "team-a"  # ISUZU CEREA VR
_TEAM_B_ID = "team-b"  # ROTHOBLAAS VOLANO TN

_ROTATION_ORDER = [
    RotationLabel.I,
    RotationLabel.VI,
    RotationLabel.V,
    RotationLabel.IV,
    RotationLabel.III,
    RotationLabel.II,
]


def _rotate(seq: list[int], n: int) -> list[int]:
    return seq[n:] + seq[:n]


def _six_for_set(
    set_number: int,
    team_id: str,
    base_order: list[int],
    rotation_shift: int,
    low_confidence_label: Optional[RotationLabel] = None,
) -> tuple[StartingSix, dict[RotationLabel, int]]:
    """Costruisce lo StartingSix di un set applicando una rotazione fissa
    dei sei titolari rispetto al set 1 — semplificazione plausibile: la
    formazione di partenza di un set spesso ruota rispetto al precedente.

    Gli id dei campi usano `starting_six_field_id` (app/volleyball/parser.py),
    la stessa convenzione della pipeline reale e dei `field_ids` emessi dal
    validator: un'unica forma di id, così il frontend naviga alle anomalie allo
    stesso modo qualunque pipeline abbia prodotto l'analisi."""

    rotated = _rotate(base_order, rotation_shift)
    labels = [
        RotationLabel.I,
        RotationLabel.II,
        RotationLabel.III,
        RotationLabel.IV,
        RotationLabel.V,
        RotationLabel.VI,
    ]
    mapping: dict[RotationLabel, int] = dict(zip(labels, rotated))

    values: dict[str, dict] = {}
    for label in labels:
        field_id = starting_six_field_id(set_number, team_id, label.value)
        player = mapping[label]
        confidence = 0.6 if label == low_confidence_label else 0.95
        values[label.value] = {
            "id": field_id,
            "value": player,
            "original_value": player,
            "confidence": confidence,
            "manually_confirmed": False,
            "source_region_id": None,
        }
    return StartingSix.model_validate(values), mapping


def _composition(total: int, parts: int, rng: random.Random) -> list[int]:
    """Divide `total` in `parts` interi positivi che sommano esattamente a
    `total` — garantisce turni di servizio con punteggio finale coerente."""

    if parts <= 1 or total <= parts:
        # fallback: un unico turno con tutti i punti, o distribuzione minima
        base = [1] * parts
        base[-1] += total - parts
        return base
    cuts = sorted(rng.sample(range(1, total), parts - 1))
    result = []
    prev = 0
    for cut in cuts + [total]:
        result.append(cut - prev)
        prev = cut
    return result


def _build_service_turns(
    set_number: int,
    first_server: str,
    final_score: tuple[int, int],
    six_a: dict[RotationLabel, int],
    six_b: dict[RotationLabel, int],
    turns_per_side: int = 5,
    low_confidence_turn_index: Optional[int] = None,
) -> list[ServiceTurn]:
    """Genera una sequenza di turni di servizio alternati, fabbricata ma
    internamente coerente: punteggio monotono, somma dei punti dei turni di
    ogni squadra = punteggio finale di quella squadra (backend §11)."""

    rng = random.Random(f"set-{set_number}-{first_server}-{final_score}")
    a_target, b_target = final_score
    a_points = _composition(a_target, turns_per_side, rng)
    b_points = _composition(b_target, turns_per_side, rng)

    turns: list[ServiceTurn] = []
    a_score = b_score = 0
    server = first_server
    rotation_index = {_TEAM_A_ID: 0, _TEAM_B_ID: 0}
    a_iter = iter(a_points)
    b_iter = iter(b_points)
    sequence = 1

    while True:
        try:
            pts = next(a_iter) if server == _TEAM_A_ID else next(b_iter)
        except StopIteration:
            break

        score_start = (a_score, b_score)
        if server == _TEAM_A_ID:
            a_score += pts
        else:
            b_score += pts
        score_end = (a_score, b_score)

        label = _ROTATION_ORDER[rotation_index[server] % len(_ROTATION_ORDER)]
        rotation_index[server] += 1
        six = six_a if server == _TEAM_A_ID else six_b
        player = six[label]

        turn_id = f"set{set_number}-turn-{sequence:03d}"
        confidence = (
            0.55 if low_confidence_turn_index == sequence else 0.9
        )

        turns.append(
            ServiceTurn(
                id=turn_id,
                sequence=sequence,
                team_id=server,
                player={
                    "id": f"{turn_id}-player",
                    "value": player,
                    "original_value": player,
                    "confidence": confidence,
                    "manually_confirmed": False,
                    "source_region_id": None,
                },
                rotation={
                    "id": f"{turn_id}-rotation",
                    "value": label.value,
                    "original_value": label.value,
                    "confidence": 0.95,
                    "manually_confirmed": False,
                    "source_region_id": None,
                },
                score_start={
                    "id": f"{turn_id}-score-start",
                    "value": list(score_start),
                    "original_value": list(score_start),
                    "confidence": 0.99,
                    "manually_confirmed": False,
                    "source_region_id": None,
                },
                score_end={
                    "id": f"{turn_id}-score-end",
                    "value": list(score_end),
                    "original_value": list(score_end),
                    "confidence": 0.99,
                    "manually_confirmed": False,
                    "source_region_id": None,
                },
                points_scored=pts,
                status=(
                    CheckStatus.WARNING
                    if confidence < 0.7
                    else CheckStatus.VALID
                ),
                source_region_ids=[],
            )
        )
        sequence += 1
        server = _TEAM_B_ID if server == _TEAM_A_ID else _TEAM_A_ID

    return turns


def _build_set(
    set_number: int,
    starting_team_id: str,
    final_score: tuple[int, int],
    cerea_order: list[int],
    rothoblaas_order: list[int],
    rotation_shift: int,
    low_confidence_label: Optional[RotationLabel] = None,
    low_confidence_turn_index: Optional[int] = None,
) -> SetData:
    six_a, mapping_a = _six_for_set(
        set_number, _TEAM_A_ID, cerea_order, rotation_shift, low_confidence_label
    )
    six_b, mapping_b = _six_for_set(
        set_number, _TEAM_B_ID, rothoblaas_order, rotation_shift
    )
    turns = _build_service_turns(
        set_number,
        starting_team_id,
        final_score,
        mapping_a,
        mapping_b,
        low_confidence_turn_index=low_confidence_turn_index,
    )

    checks: list[ValidationCheck] = [
        ValidationCheck(
            id=f"set{set_number}-rotation-order",
            label="Ordine delle rotazioni",
            status=CheckStatus.VALID,
            message=None,
            field_ids=[],
        )
    ]
    status = CheckStatus.VALID
    if low_confidence_label is not None or low_confidence_turn_index is not None:
        low_field_ids = []
        if low_confidence_label is not None:
            low_field_ids.append(
                starting_six_field_id(set_number, _TEAM_A_ID, low_confidence_label.value)
            )
        if low_confidence_turn_index is not None:
            low_field_ids.append(f"set{set_number}-turn-{low_confidence_turn_index:03d}-player")
        checks.append(
            ValidationCheck(
                id=f"set{set_number}-low-confidence",
                label="Dati incerti",
                status=CheckStatus.WARNING,
                message=f"{len(low_field_ids)} valori richiedono verifica",
                field_ids=low_field_ids,
            )
        )
        status = CheckStatus.WARNING

    return SetData(
        number=set_number,
        starting_team_id=starting_team_id,
        team_a_starting_six=six_a,
        team_b_starting_six=six_b,
        service_turns=turns,
        final_score=final_score,
        validation=ValidationResult(status=status, checks=checks),
    )


def build_canned_analysis(analysis_id: str) -> Analysis:
    """Analysis CANNED completa (backend §28, §44) — **dati fabbricati**.

    Non è più la pipeline di default: quella è `run_real_pipeline`. Resta come
    fallback per i casi in cui l'estrazione reale non è disponibile, e chi la
    usa DEVE marcare il risultato (vedi `AnalysisService._canned_fallback`):
    questi numeri non provengono da nessun PDF.
    """

    cerea_order_set1 = [2, 5, 3, 8, 14, 9]  # I, II, III, IV, V, VI
    rothoblaas_order_set1 = [14, 9, 3, 4, 15, 17]

    sets = [
        _build_set(
            1,
            starting_team_id=_TEAM_B_ID,
            final_score=(25, 27),
            cerea_order=cerea_order_set1,
            rothoblaas_order=rothoblaas_order_set1,
            rotation_shift=0,
            low_confidence_label=RotationLabel.V,
        ),
        _build_set(
            2,
            starting_team_id=_TEAM_A_ID,
            final_score=(25, 20),
            cerea_order=cerea_order_set1,
            rothoblaas_order=rothoblaas_order_set1,
            rotation_shift=1,
            low_confidence_turn_index=3,
        ),
        _build_set(
            3,
            starting_team_id=_TEAM_B_ID,
            final_score=(25, 22),
            cerea_order=cerea_order_set1,
            rothoblaas_order=rothoblaas_order_set1,
            rotation_shift=2,
        ),
        _build_set(
            4,
            starting_team_id=_TEAM_A_ID,
            final_score=(25, 23),
            cerea_order=cerea_order_set1,
            rothoblaas_order=rothoblaas_order_set1,
            rotation_shift=3,
        ),
    ]

    overall_validation = CheckStatus.VALID
    all_checks: list[ValidationCheck] = []
    for s in sets:
        all_checks.extend(s.validation.checks)
        if s.validation.status != CheckStatus.VALID:
            overall_validation = CheckStatus.WARNING

    match = MatchInfo(
        competition="Serie B1",
        match_number="8477",
        date="2025-12-20",
        time="21:00",
        venue="CEREA",
        team_a=Team(id=_TEAM_A_ID, name="ISUZU CEREA VR"),
        team_b=Team(id=_TEAM_B_ID, name="ROTHOBLAAS VOLANO TN"),
        final_result=(3, 1),
    )

    return Analysis(
        id=analysis_id,
        status=AnalysisGlobalStatus.READY,
        overall_validation=overall_validation,
        match=match,
        sets=sets,
        source_regions=[],
        validation=ValidationResult(status=overall_validation, checks=all_checks),
    )


# --------------------------------------------------------------------------
# ExtractedValue lookup/mutation helpers (per PATCH e reset-corrections).
# --------------------------------------------------------------------------


def _is_extracted_value_node(node: object) -> bool:
    return (
        isinstance(node, dict)
        and "id" in node
        and "value" in node
        and "original_value" in node
        and "manually_confirmed" in node
    )


def _find_and_update_field(data: object, field_id: str, new_value: object) -> bool:
    """Cerca recursivamente un nodo ExtractedValue con id == field_id e ne
    aggiorna `value` impostando `manually_confirmed = True`. Ritorna True se
    trovato. `original_value` non viene mai toccato qui (backend §8/§13)."""

    if isinstance(data, dict):
        if _is_extracted_value_node(data) and data["id"] == field_id:
            data["value"] = new_value
            data["manually_confirmed"] = True
            return True
        for value in data.values():
            if _find_and_update_field(value, field_id, new_value):
                return True
        return False
    if isinstance(data, list):
        for item in data:
            if _find_and_update_field(item, field_id, new_value):
                return True
        return False
    return False


def _reset_all_fields(data: object) -> None:
    """Ripristina value=original_value e manually_confirmed=False su ogni
    nodo ExtractedValue trovato, ricorsivamente (backend §15)."""

    if isinstance(data, dict):
        if _is_extracted_value_node(data):
            data["value"] = data["original_value"]
            data["manually_confirmed"] = False
        for value in data.values():
            _reset_all_fields(value)
    elif isinstance(data, list):
        for item in data:
            _reset_all_fields(item)


# --------------------------------------------------------------------------
# Validazione upload (backend §5, §40 minimo: MIME/estensione + signature).
# --------------------------------------------------------------------------

_PDF_MAGIC = b"%PDF-"


def _validate_upload(filename: Optional[str], content_type: Optional[str], content: bytes) -> None:
    if not filename or not filename.lower().endswith(".pdf"):
        raise InvalidFileError(
            "Il file deve avere estensione .pdf.",
            details={"filename": filename},
        )
    if content_type not in (None, "application/pdf", "application/octet-stream"):
        raise InvalidFileError(
            "Content-Type non valido per un PDF.",
            details={"content_type": content_type},
        )
    if not content:
        raise InvalidFileError("Il file è vuoto.")
    if not content.startswith(_PDF_MAGIC):
        raise UnsupportedPdfError(
            "Il documento non può essere interpretato: signature PDF assente."
        )


# --------------------------------------------------------------------------
# AnalysisService
# --------------------------------------------------------------------------


class AnalysisService:
    def __init__(self, repository: AnalysisRepository, storage_dir: Path) -> None:
        self._repository = repository
        self._storage_dir = storage_dir

    # -- creazione --------------------------------------------------------

    async def create_analysis(
        self,
        *,
        filename: Optional[str],
        content_type: Optional[str],
        content: bytes,
        background_tasks: BackgroundTasks,
    ) -> CreateAnalysisResponse:
        _validate_upload(filename, content_type, content)

        analysis_id = str(uuid.uuid4())
        analysis_dir = self._storage_dir / analysis_id
        analysis_dir.mkdir(parents=True, exist_ok=True)
        pdf_path = analysis_dir / "original.pdf"
        pdf_path.write_bytes(content)

        record = self._repository.create(
            analysis_id=analysis_id,
            original_filename=filename or "document.pdf",
            pdf_path=str(pdf_path),
            initial_analysis_json=_skeleton_analysis(analysis_id),
        )

        background_tasks.add_task(self._run_pipeline, analysis_id)

        return CreateAnalysisResponse(
            analysis_id=analysis_id, status=AnalysisGlobalStatus(record.status)
        )

    # -- pipeline -----------------------------------------------------------

    async def _run_pipeline(self, analysis_id: str) -> None:
        try:
            await self._run_pipeline_steps(analysis_id)
        except Exception as exc:  # noqa: BLE001 - qualunque errore ⇒ stato terminale
            # Rete di sicurezza: se qualcosa qui alza (errore di storage, bug non
            # previsto, ecc.) l'analisi NON deve restare bloccata su PROCESSING —
            # sarebbe uno spinner infinito lato frontend, che fa polling finché lo
            # stato non è terminale. La marchiamo FAILED con un messaggio leggibile.
            # NB: un OOM (SIGKILL) non è catturabile qui — quel caso è coperto
            # dalla riconciliazione all'avvio (`fail_unfinished_analyses`).
            _logger.error(
                "pipeline fallita in modo non gestito, analisi marcata FAILED",
                exc_info=True,
                extra={"analysis_id": analysis_id, "error_type": type(exc).__name__},
            )
            self._mark_failed(analysis_id, f"{type(exc).__name__}: {exc}")

    async def _run_pipeline_steps(self, analysis_id: str) -> None:
        total_steps = len(_STEP_ORDER)
        analysis: Optional[Analysis] = None
        for index in range(total_steps):
            progress_processing = int(round((index / total_steps) * 100))
            self._repository.update_progress(
                analysis_id,
                status=AnalysisGlobalStatus.PROCESSING.value,
                progress=progress_processing,
                current_step=_STEP_ORDER[index].value,
                steps=_steps_snapshot(completed_up_to=index, processing_index=index),
            )
            if _STEP_ORDER[index] is ProcessingStepId.READ_DOCUMENT:
                # Il lavoro vero (rendering, OCR, parsing) è sincrono e può
                # durare decine di secondi sul percorso raster: va fuori dal
                # loop asyncio, altrimenti il polling di `GET .../status`
                # resterebbe bloccato per tutta la durata dell'estrazione.
                analysis = await asyncio.to_thread(self._produce_analysis, analysis_id)
            else:
                await asyncio.sleep(_STEP_DELAY_SECONDS)

            progress_completed = int(round(((index + 1) / total_steps) * 100))
            self._repository.update_progress(
                analysis_id,
                status=AnalysisGlobalStatus.PROCESSING.value,
                progress=progress_completed,
                current_step=_STEP_ORDER[index].value,
                steps=_steps_snapshot(completed_up_to=index + 1, processing_index=None),
            )

        if analysis is None:  # pragma: no cover - difensivo
            analysis = self._canned_fallback(analysis_id, "pipeline non eseguita")
        self._repository.save_result(
            analysis_id,
            status=AnalysisGlobalStatus.READY.value,
            analysis_json=analysis.model_dump(mode="json"),
        )
        self._repository.update_progress(
            analysis_id,
            status=AnalysisGlobalStatus.READY.value,
            progress=100,
            current_step=None,
            steps=_steps_snapshot(completed_up_to=total_steps, processing_index=None),
        )

    def _mark_failed(self, analysis_id: str, reason: str) -> None:
        """Porta l'analisi allo stato terminale FAILED con un errore leggibile.

        Conserva progress/steps già raggiunti (utile a chi guarda a che punto si
        è interrotta) ma azzera `current_step` e imposta un `ErrorDetail`
        `ANALYSIS_FAILED`, che il frontend mostra nel riquadro d'errore."""

        record = self._repository.get(analysis_id)
        self._repository.update_progress(
            analysis_id,
            status=AnalysisGlobalStatus.FAILED.value,
            progress=record.progress if record else 0,
            current_step=None,
            steps=record.steps if record else _steps_snapshot(0, None),
            error={
                "code": ErrorCode.ANALYSIS_FAILED.value,
                "message": (
                    "L'analisi del referto non è riuscita. Riprova; se il problema "
                    "persiste, il documento potrebbe non essere leggibile."
                ),
                "details": {"reason": reason},
            },
        )

    def fail_unfinished_analyses(self) -> int:
        """Marca FAILED le analisi rimaste non terminali da un processo morto.

        I task di estrazione girano in-process (FastAPI BackgroundTasks): se il
        processo viene ucciso mentre elabora — tipicamente un OOM su un referto
        scansionato pesante — lo stato resta congelato su UPLOADED/PROCESSING e
        il polling del frontend non finirebbe mai. All'avvio nessun task del
        processo precedente è più vivo, quindi ogni analisi non terminale è, di
        fatto, fallita: la si chiude con un errore esplicito così che il client
        smetta di girare a vuoto e possa proporre "Riprova"."""

        error = {
            "code": ErrorCode.ANALYSIS_FAILED.value,
            "message": (
                "L'elaborazione si è interrotta perché il servizio è stato "
                "riavviato durante l'analisi. Riprova l'analisi del referto."
            ),
            "details": {"reason": "process-restart"},
        }
        count = self._repository.fail_unfinished(error=error)
        if count:
            _logger.warning(
                "analisi non terminate marcate FAILED all'avvio",
                extra={"count": count},
            )
        return count

    # -- scelta fra pipeline reale e fallback canned -------------------------

    def _produce_analysis(self, analysis_id: str) -> Analysis:
        """Esegue l'estrazione reale sul PDF caricato, con fallback tracciato.

        Un PDF non supportato non deve far crashare l'app, ma non deve nemmeno
        ricadere in silenzio su dati fabbricati: ogni ripiego passa da
        `_canned_fallback`, che logga e marca il risultato.
        """

        if not real_pipeline_enabled():
            return self._canned_fallback(
                analysis_id, f"pipeline reale disabilitata da {REAL_PIPELINE_ENV_VAR}"
            )

        record = self._repository.get(analysis_id)
        if record is None or not record.pdf_path:
            return self._canned_fallback(analysis_id, "PDF originale non disponibile")

        try:
            result = run_real_pipeline(record.pdf_path, analysis_id)
        except Exception as exc:  # noqa: BLE001 - qualunque errore ⇒ fallback tracciato
            _logger.error(
                "estrazione reale fallita, ripiego su dati simulati",
                exc_info=True,
                extra={
                    "analysis_id": analysis_id,
                    "pdf_path": record.pdf_path,
                    "error_type": type(exc).__name__,
                },
            )
            return self._canned_fallback(analysis_id, f"{type(exc).__name__}: {exc}")

        _logger.info(
            "analisi prodotta dalla pipeline reale",
            extra={
                "analysis_id": analysis_id,
                "method": result.method.value,
                "sets": [s.number for s in result.analysis.sets],
                "overall_validation": result.analysis.overall_validation.value
                if result.analysis.overall_validation
                else None,
                "diagnostics": result.diagnostics,
            },
        )
        return result.analysis

    def _canned_fallback(self, analysis_id: str, reason: str) -> Analysis:
        """Analysis canned marcata come NON proveniente dal PDF caricato."""

        _logger.error(
            "risultato canned servito al posto dell'estrazione reale",
            extra={"analysis_id": analysis_id, "reason": reason},
        )
        analysis = build_canned_analysis(analysis_id)
        analysis.validation.checks.insert(0, fallback_check(reason))
        analysis.validation.status = CheckStatus.INVALID
        analysis.overall_validation = CheckStatus.INVALID
        return analysis

    # -- lettura ------------------------------------------------------------

    def _get_record_or_raise(self, analysis_id: str) -> AnalysisRecord:
        record = self._repository.get(analysis_id)
        if record is None:
            raise AnalysisNotFoundError(f"Analysis '{analysis_id}' non trovata.")
        return record

    def get_status(self, analysis_id: str) -> AnalysisStatusResponse:
        record = self._get_record_or_raise(analysis_id)
        steps = [ProcessingStep.model_validate(s) for s in record.steps]
        return AnalysisStatusResponse(
            analysis_id=record.id,
            status=AnalysisGlobalStatus(record.status),
            progress=record.progress,
            current_step=ProcessingStepId(record.current_step)
            if record.current_step
            else None,
            steps=steps,
            error=record.error,
        )

    def get_analysis(self, analysis_id: str) -> Analysis:
        record = self._get_record_or_raise(analysis_id)
        return Analysis.model_validate(record.analysis_json or {})

    def get_source_pdf_path(self, analysis_id: str) -> Path:
        """Path assoluto e verificato del PDF originale caricato per
        l'analisi (usato da `GET /analyses/{id}/source-pdf`); il
        contenimento nella storage dir e l'esistenza su disco sono
        validati da `app.core.security.resolve_pdf_within_storage`."""
        record = self._get_record_or_raise(analysis_id)
        return resolve_pdf_within_storage(self._storage_dir, record.pdf_path)

    # -- correzioni manuali ---------------------------------------------------

    def patch_field(self, analysis_id: str, field_id: str, value: object) -> Analysis:
        record = self._get_record_or_raise(analysis_id)
        data = record.analysis_json or {}
        updated = _find_and_update_field(data, field_id, value)
        if not updated:
            raise InvalidFieldValueError(
                f"Campo '{field_id}' non trovato.", details={"field_id": field_id}
            )
        try:
            analysis = Analysis.model_validate(data)
        except Exception as exc:  # pydantic ValidationError o simili
            raise InvalidFieldValueError(
                f"Valore non valido per il campo '{field_id}'.",
                details={"field_id": field_id, "reason": str(exc)},
            ) from exc

        self._repository.update_analysis_json(analysis_id, analysis.model_dump(mode="json"))
        return analysis

    def reset_corrections(self, analysis_id: str) -> Analysis:
        record = self._get_record_or_raise(analysis_id)
        data = record.analysis_json or {}
        _reset_all_fields(data)
        analysis = Analysis.model_validate(data)
        self._repository.update_analysis_json(analysis_id, analysis.model_dump(mode="json"))
        return analysis

    # -- rianalisi ------------------------------------------------------------

    def reanalyze(self, analysis_id: str, background_tasks: BackgroundTasks) -> None:
        self._get_record_or_raise(analysis_id)  # 404 se l'id non esiste
        self._repository.reset_for_reanalysis(
            analysis_id, initial_analysis_json=_skeleton_analysis(analysis_id)
        )
        background_tasks.add_task(self._run_pipeline, analysis_id)

    # -- export (placeholder — backend §30/§31, implementazione reale altrove) --

    def export_xlsx(self, analysis_id: str) -> bytes:
        from io import BytesIO

        from openpyxl import Workbook

        analysis = self.get_analysis(analysis_id)

        wb = Workbook()
        match_sheet = wb.active
        match_sheet.title = "Match"
        match_sheet.append(["Competition", analysis.match.competition])
        match_sheet.append(["Match number", analysis.match.match_number])
        match_sheet.append(["Date", analysis.match.date])
        match_sheet.append(["Team A", analysis.match.team_a.name])
        match_sheet.append(["Team B", analysis.match.team_b.name])
        match_sheet.append(["Final result", str(list(analysis.match.final_result))])

        six_sheet = wb.create_sheet("Starting Six")
        six_sheet.append(["Set", "Team", "I", "II", "III", "IV", "V", "VI"])
        for s in analysis.sets:
            six_sheet.append(
                [s.number, analysis.match.team_a.name]
                + [getattr(s.team_a_starting_six, label).value for label in ("I", "II", "III", "IV", "V", "VI")]
            )
            six_sheet.append(
                [s.number, analysis.match.team_b.name]
                + [getattr(s.team_b_starting_six, label).value for label in ("I", "II", "III", "IV", "V", "VI")]
            )

        turns_sheet = wb.create_sheet("Service Turns")
        turns_sheet.append(
            [
                "Set",
                "Sequence",
                "Team",
                "Player",
                "Rotation",
                "Score Start A",
                "Score Start B",
                "Score End A",
                "Score End B",
                "Points Scored",
                "Confidence",
                "Status",
            ]
        )
        for s in analysis.sets:
            for turn in s.service_turns:
                turns_sheet.append(
                    [
                        s.number,
                        turn.sequence,
                        turn.team_id,
                        turn.player.value,
                        turn.rotation.value.value if turn.rotation.value else None,
                        turn.score_start.value[0],
                        turn.score_start.value[1],
                        turn.score_end.value[0],
                        turn.score_end.value[1],
                        turn.points_scored,
                        turn.player.confidence,
                        turn.status.value,
                    ]
                )

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def export_csv(self, analysis_id: str, dataset: str) -> str:
        import csv
        from io import StringIO

        analysis = self.get_analysis(analysis_id)
        buffer = StringIO()
        writer = csv.writer(buffer)

        if dataset == "starting-six":
            writer.writerow(["Set", "Team", "I", "II", "III", "IV", "V", "VI"])
            for s in analysis.sets:
                writer.writerow(
                    [s.number, analysis.match.team_a.name]
                    + [getattr(s.team_a_starting_six, label).value for label in ("I", "II", "III", "IV", "V", "VI")]
                )
                writer.writerow(
                    [s.number, analysis.match.team_b.name]
                    + [getattr(s.team_b_starting_six, label).value for label in ("I", "II", "III", "IV", "V", "VI")]
                )
        elif dataset == "service-turns":
            writer.writerow(
                [
                    "Set",
                    "Sequence",
                    "Team",
                    "Player",
                    "Rotation",
                    "Score Start A",
                    "Score Start B",
                    "Score End A",
                    "Score End B",
                    "Points Scored",
                    "Confidence",
                    "Status",
                ]
            )
            for s in analysis.sets:
                for turn in s.service_turns:
                    writer.writerow(
                        [
                            s.number,
                            turn.sequence,
                            turn.team_id,
                            turn.player.value,
                            turn.rotation.value.value if turn.rotation.value else None,
                            turn.score_start.value[0],
                            turn.score_start.value[1],
                            turn.score_end.value[0],
                            turn.score_end.value[1],
                            turn.points_scored,
                            turn.player.confidence,
                            turn.status.value,
                        ]
                    )
        else:
            raise ExportFailedError(
                f"Dataset export non supportato: '{dataset}'.",
                details={"dataset": dataset},
            )

        return buffer.getvalue()
