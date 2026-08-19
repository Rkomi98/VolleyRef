"""Generazione dell'export Excel (.xlsx) di un'Analysis.

Specifica di riferimento: 02_volleyref_backend_prompt.md §30. Il workbook
prodotto ha tre sheet:

- ``Match``: informazioni generali sulla partita e punteggio di ogni set.
- ``Starting Six``: sestetti di partenza per set/squadra (colonne
  ``Set, Team, I, II, III, IV, V, VI``, esattamente come richiesto dalla
  spec).
- ``Service Turns``: turni di servizio (colonne ``Set, Sequence, Team,
  Player, Rotation, Score Start A, Score Start B, Score End A, Score End B,
  Points Scored, Confidence, Status``).

Usa sempre i valori correnti dell'``Analysis`` (``ExtractedValue.value``),
comprese le correzioni manuali già applicate — non ``original_value``
(§30: "Usa i valori correnti, comprese le correzioni manuali").
"""

from __future__ import annotations

from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.models.analysis import Analysis, AnalysisGlobalStatus
from app.models.match import ServiceTurn
from app.services.analysis_service import ExportFailedError

_HEADER_FONT = Font(bold=True)

_STARTING_SIX_LABELS = ("I", "II", "III", "IV", "V", "VI")

_SERVICE_TURNS_HEADERS = [
    "Set",
    "Sequence",
    "Team",
    "Player",
    "Rotation",
    "Score Start A",
    "Score Start B",
    "Score End A",
    "Score End B",
    "Points Scored",
    "Confidence",
    "Status",
]


def _require_ready(analysis: Analysis) -> None:
    """L'export ha senso solo su un'Analysis completa (backend §30/§31).

    Un'Analysis non ancora READY (UPLOADED/PROCESSING/FAILED) non ha dati
    affidabili da esportare: viene rifiutata con lo stesso ErrorCode già
    usato per gli altri errori di export (EXPORT_FAILED, backend §34).
    """

    if analysis.status != AnalysisGlobalStatus.READY:
        raise ExportFailedError(
            "L'analisi non è pronta per l'export: è richiesto lo stato READY.",
            details={"analysis_id": analysis.id, "status": analysis.status.value},
        )


def _write_header(sheet: Worksheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[sheet.max_row]:
        cell.font = _HEADER_FONT


def _set_column_widths(sheet: Worksheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        column_letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[column_letter].width = width


def _rotation_value(turn: ServiceTurn) -> Optional[str]:
    rotation = turn.rotation.value
    return rotation.value if rotation is not None else None


def _build_match_sheet(sheet: Worksheet, analysis: Analysis) -> None:
    sheet.title = "Match"
    match = analysis.match

    general_info = [
        ("Competition", match.competition),
        ("Match number", match.match_number),
        ("Date", match.date),
        ("Time", match.time),
        ("Venue", match.venue),
        ("Team A", match.team_a.name),
        ("Team B", match.team_b.name),
        ("Final result", f"{match.final_result[0]}-{match.final_result[1]}"),
        (
            "Overall validation",
            analysis.overall_validation.value if analysis.overall_validation else None,
        ),
    ]
    for label, value in general_info:
        row = sheet.max_row + 1
        label_cell = sheet.cell(row=row, column=1, value=label)
        label_cell.font = _HEADER_FONT
        sheet.cell(row=row, column=2, value=value)

    sheet.append([])  # separatore visivo tra informazioni generali e set

    _write_header(sheet, ["Set", "Team A Score", "Team B Score", "Validation"])
    for set_data in analysis.sets:
        sheet.append(
            [
                set_data.number,
                set_data.final_score[0],
                set_data.final_score[1],
                set_data.validation.status.value,
            ]
        )

    _set_column_widths(sheet, [20, 24, 16, 16])


def _build_starting_six_sheet(workbook: Workbook, analysis: Analysis) -> None:
    sheet = workbook.create_sheet("Starting Six")
    _write_header(sheet, ["Set", "Team", *_STARTING_SIX_LABELS])

    for set_data in analysis.sets:
        for team_name, starting_six in (
            (analysis.match.team_a.name, set_data.team_a_starting_six),
            (analysis.match.team_b.name, set_data.team_b_starting_six),
        ):
            sheet.append(
                [set_data.number, team_name]
                + [getattr(starting_six, label).value for label in _STARTING_SIX_LABELS]
            )

    _set_column_widths(sheet, [8, 26, 6, 6, 6, 6, 6, 6])


def _build_service_turns_sheet(workbook: Workbook, analysis: Analysis) -> None:
    sheet = workbook.create_sheet("Service Turns")
    _write_header(sheet, _SERVICE_TURNS_HEADERS)

    for set_data in analysis.sets:
        for turn in set_data.service_turns:
            sheet.append(
                [
                    set_data.number,
                    turn.sequence,
                    turn.team_id,
                    turn.player.value,
                    _rotation_value(turn),
                    turn.score_start.value[0],
                    turn.score_start.value[1],
                    turn.score_end.value[0],
                    turn.score_end.value[1],
                    turn.points_scored,
                    turn.player.confidence,
                    turn.status.value,
                ]
            )

    _set_column_widths(sheet, [6, 10, 20, 8, 10, 14, 14, 14, 14, 14, 12, 10])


def build_xlsx(analysis: Analysis) -> bytes:
    """Costruisce il workbook .xlsx completo per un'``Analysis`` (§30).

    Ritorna i byte del file .xlsx pronti per essere restituiti come
    download HTTP (``Content-Disposition: attachment``). Alza
    ``ExportFailedError`` se l'analisi non è nello stato READY.
    """

    _require_ready(analysis)

    workbook = Workbook()
    match_sheet = workbook.active  # primo sheet creato di default da Workbook()
    _build_match_sheet(match_sheet, analysis)
    _build_starting_six_sheet(workbook, analysis)
    _build_service_turns_sheet(workbook, analysis)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
