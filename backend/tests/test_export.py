"""Test per gli export xlsx/csv di un'Analysis (backend §30-§31).

Costruisce un'Analysis READY minimale direttamente con i modelli Pydantic
(non passa dalla pipeline mock), poi verifica che ``app.export.xlsx.build_xlsx``
e ``app.export.csv.build_csv`` producano file validi e apribili, con lo
schema di sheet/colonne richiesto dalla spec. Verifica anche che un'Analysis
non ancora READY venga rifiutata con ``ExportFailedError``.
"""

from __future__ import annotations

from io import BytesIO

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.export.csv import build_csv
from app.export.xlsx import build_xlsx
from app.models.analysis import Analysis, AnalysisGlobalStatus
from app.models.common import CheckStatus
from app.models.match import (
    MatchInfo,
    RotationLabel,
    ServiceTurn,
    SetData,
    StartingSix,
    Team,
    ValidationResult,
)
from app.services.analysis_service import ExportFailedError

_STARTING_SIX_LABELS = ["I", "II", "III", "IV", "V", "VI"]

_SERVICE_TURNS_COLUMNS = [
    "Set",
    "Sequence",
    "Team",
    "Player",
    "Rotation",
    "Score Start A",
    "Score Start B",
    "Score End A",
    "Score End B",
    "Points Scored",
    "Confidence",
    "Status",
]


def _extracted(field_id: str, value: object, confidence: float = 0.9) -> dict:
    return {
        "id": field_id,
        "value": value,
        "original_value": value,
        "confidence": confidence,
        "manually_confirmed": False,
        "source_region_id": None,
    }


def _starting_six(prefix: str, players: list[int]) -> StartingSix:
    return StartingSix.model_validate(
        {
            label: _extracted(f"{prefix}-{label}", player)
            for label, player in zip(_STARTING_SIX_LABELS, players)
        }
    )


def _service_turn(
    sequence: int,
    team_id: str,
    player: int,
    rotation: RotationLabel,
    score_start: tuple[int, int],
    score_end: tuple[int, int],
) -> ServiceTurn:
    turn_id = f"turn-{sequence}"
    points_scored = (score_end[0] + score_end[1]) - (score_start[0] + score_start[1])
    return ServiceTurn(
        id=turn_id,
        sequence=sequence,
        team_id=team_id,
        player=_extracted(f"{turn_id}-player", player),
        rotation=_extracted(f"{turn_id}-rotation", rotation.value),
        score_start=_extracted(f"{turn_id}-score-start", list(score_start)),
        score_end=_extracted(f"{turn_id}-score-end", list(score_end)),
        points_scored=points_scored,
        status=CheckStatus.VALID,
        source_region_ids=[],
    )


def _build_test_analysis(
    status: AnalysisGlobalStatus = AnalysisGlobalStatus.READY,
) -> Analysis:
    """Analysis in memoria con un solo set, due sestetti e tre turni di
    servizio — sufficiente a esercitare tutte le colonne dell'export senza
    passare dalla pipeline mock del servizio."""

    team_a = Team(id="team-a", name="Team Alpha")
    team_b = Team(id="team-b", name="Team Beta")

    six_a = _starting_six("set1-a", [1, 2, 3, 4, 5, 6])
    six_b = _starting_six("set1-b", [11, 12, 13, 14, 15, 16])

    turns = [
        _service_turn(1, "team-a", 1, RotationLabel.I, (0, 0), (1, 0)),
        _service_turn(2, "team-b", 11, RotationLabel.I, (1, 0), (1, 1)),
        _service_turn(3, "team-a", 2, RotationLabel.II, (1, 1), (2, 1)),
    ]

    set1 = SetData(
        number=1,
        starting_team_id="team-a",
        team_a_starting_six=six_a,
        team_b_starting_six=six_b,
        service_turns=turns,
        final_score=(25, 20),
        validation=ValidationResult(status=CheckStatus.VALID, checks=[]),
    )

    match = MatchInfo(
        competition="Serie B1",
        match_number="1234",
        date="2026-01-10",
        time="20:30",
        venue="Palestra Test",
        team_a=team_a,
        team_b=team_b,
        final_result=(1, 0),
    )

    return Analysis(
        id="analysis-test",
        status=status,
        overall_validation=CheckStatus.VALID,
        match=match,
        sets=[set1],
        source_regions=[],
        validation=ValidationResult(status=CheckStatus.VALID, checks=[]),
    )


class TestBuildXlsx:
    def test_generates_workbook_with_expected_sheets_and_values(self) -> None:
        analysis = _build_test_analysis()
        content = build_xlsx(analysis)

        workbook = load_workbook(BytesIO(content))
        assert workbook.sheetnames == ["Match", "Starting Six", "Service Turns"]

        match_sheet = workbook["Match"]
        match_rows = {
            row[0].value: row[1].value
            for row in match_sheet.iter_rows(max_row=9)
            if row[0].value
        }
        assert match_rows["Competition"] == "Serie B1"
        assert match_rows["Team A"] == "Team Alpha"
        assert match_rows["Team B"] == "Team Beta"
        assert match_rows["Final result"] == "1-0"

        six_sheet = workbook["Starting Six"]
        header = [cell.value for cell in six_sheet[1]]
        assert header == ["Set", "Team", "I", "II", "III", "IV", "V", "VI"]
        row_a = [cell.value for cell in six_sheet[2]]
        row_b = [cell.value for cell in six_sheet[3]]
        assert row_a == [1, "Team Alpha", 1, 2, 3, 4, 5, 6]
        assert row_b == [1, "Team Beta", 11, 12, 13, 14, 15, 16]

        turns_sheet = workbook["Service Turns"]
        turns_header = [cell.value for cell in turns_sheet[1]]
        assert turns_header == _SERVICE_TURNS_COLUMNS
        assert turns_sheet.max_row == 1 + len(analysis.sets[0].service_turns)
        first_turn_row = [cell.value for cell in turns_sheet[2]]
        assert first_turn_row == [1, 1, "team-a", 1, "I", 0, 0, 1, 0, 1, 0.9, "VALID"]

    def test_header_row_is_bold(self) -> None:
        analysis = _build_test_analysis()
        workbook = load_workbook(BytesIO(build_xlsx(analysis)))
        six_sheet = workbook["Starting Six"]
        assert all(cell.font.bold for cell in six_sheet[1])

    @pytest.mark.parametrize(
        "status",
        [AnalysisGlobalStatus.UPLOADED, AnalysisGlobalStatus.PROCESSING, AnalysisGlobalStatus.FAILED],
    )
    def test_raises_export_failed_when_analysis_not_ready(
        self, status: AnalysisGlobalStatus
    ) -> None:
        analysis = _build_test_analysis(status=status)
        with pytest.raises(ExportFailedError):
            build_xlsx(analysis)


class TestBuildCsv:
    def test_starting_six_dataset_round_trips_through_pandas(self) -> None:
        analysis = _build_test_analysis()
        content = build_csv(analysis, "starting-six")

        frame = pd.read_csv(BytesIO(content.encode("utf-8")))
        assert list(frame.columns) == ["Set", "Team", "I", "II", "III", "IV", "V", "VI"]
        assert len(frame) == 2

        row_a = frame[frame["Team"] == "Team Alpha"].iloc[0]
        assert int(row_a["I"]) == 1
        assert int(row_a["VI"]) == 6

    def test_service_turns_dataset_round_trips_through_pandas(self) -> None:
        analysis = _build_test_analysis()
        content = build_csv(analysis, "service-turns")

        frame = pd.read_csv(BytesIO(content.encode("utf-8")))
        assert list(frame.columns) == _SERVICE_TURNS_COLUMNS
        assert len(frame) == 3
        assert frame.iloc[0]["Player"] == 1
        assert frame.iloc[0]["Rotation"] == "I"
        assert frame.iloc[0]["Status"] == "VALID"

    def test_unknown_dataset_raises_export_failed(self) -> None:
        analysis = _build_test_analysis()
        with pytest.raises(ExportFailedError):
            build_csv(analysis, "bogus")

    def test_raises_export_failed_when_analysis_not_ready(self) -> None:
        analysis = _build_test_analysis(status=AnalysisGlobalStatus.UPLOADED)
        with pytest.raises(ExportFailedError):
            build_csv(analysis, "starting-six")
